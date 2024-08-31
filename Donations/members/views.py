from django.shortcuts import render, redirect
from django.core.mail import EmailMessage 
from openpyxl import Workbook, load_workbook
from .models import Donators
from .models import inNeed
from .models import Post

from django.conf import settings
 
import os

def form_view(request):
    return render(request, 'index.html')

def donator(request):
    return render(request, 'Donator.html')

def in_need(request):
    return render(request, 'inNeed.html')
 

def submit(request):
    if request.method == 'POST':
        name = request.POST['name']
        email = request.POST['email']
        number=request.POST['number']
        message = request.POST['message']
      
        form_data = Donators(name=name, email=email, message=message)
        form_data.save()
        
     
        file_path = 'Donator.xlsx'
        if os.path.exists(file_path):
             workbook = load_workbook(file_path)
             sheet = workbook.active
        else:
          workbook = Workbook()
          sheet = workbook.active
          sheet.append(['الاسم', 'الايميل','رقم الهاتف', 'التبرعات'])
        
        # Append the new data
        sheet.append([name, email,number, message])
        workbook.save(file_path)
        
 
        send_email(file_path)

    return redirect('form')


def submit1(request):
    if request.method == 'POST':
        name = request.POST['name']
        email = request.POST['email']
        number = request.POST['number']
        message = request.POST['message']
        
     
        form_data = inNeed(name=name, email=email, message=message)
        form_data.save()
        
 
        file_path = 'inNeed.xlsx'
        if os.path.exists(file_path):
             workbook = load_workbook(file_path)
             sheet = workbook.active
        else:
          workbook = Workbook()
          sheet = workbook.active
          sheet.append(['الاسم', 'الايميل','رقم الهاتف', 'الاحتياجات'])
        
      
        sheet.append([name, email,number, message])
        workbook.save(file_path)
        
 
        send_email(file_path)

    return redirect('form')


def send_email(file_path):
    admin_email = 'donations4keraza@gmail.com'
    from_email = 'donations4keraza@gmail.com'
    subject = 'Updated Data'
    body = 'Please find the updated data attached.'

    email = EmailMessage(
        subject,
        body,
        from_email,
        [admin_email]
    )
    email.attach_file(file_path)
    email.send()

def blogs(request):
    posts = Post.objects.all()
    return render(request, 'blogs.html', {'posts': posts})

def post(request, pk):
    posts=Post.objects.get(id=pk)
    return render(request, 'posts.html', {'posts':posts} )

def posting(request):
   if request.method =='POST':
     title=request.POST['title']
     field=request.POST['field']
     Post_data = Post(title=title, body=field)
     Post_data.save()
     return redirect('blogs' )
   else:
      return render(request,'posting.html')
